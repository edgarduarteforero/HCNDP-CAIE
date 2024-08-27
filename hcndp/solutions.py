# -*- coding: utf-8 -*-
"""
Created on Wed Dec 20 17:58:33 2023
@author: edgar
"""
from hcndp import local_search
from hcndp import tabu_search
from hcndp import vnd
from hcndp import gvns
from hcndp import initial_solution

import textwrap
import pandas as pd
import numpy as np
from hcndp import landscape
import pyomo.environ as pyo
import os
#from idaes.core.util.model_statistics import degrees_of_freedom
#from idaes.core.util.model_statistics import report_statistics
#import xlsxwriter
#import os
#import openpyxl
import xlsxwriter
import openpyxl
from hcndp import kpi
from hcndp import network
import copy
from hcndp.data_functions import indices
from itertools import product
from hcndp import models
from hcndp import read_data
import shutil
from hcndp import data_functions
from hcndp import solutions
import time


def menu_solutions(network_original, problems_dict):
    
    while True:
        print("\n----------------------------------------------------------")
        print(f"Vamos a agregar soluciones a la red {network_original.name}.")
        print("menu_solutions@solutions.py")
        print("----------------------------------------------------------\n")
        print("Selecciona una opción:")
        print("1. Cargar tu propia solución")
        print("2. Obtener soluciones mono-objetivo")
        print("4. Indicadores (KPI) y gráficos de soluciones")
        print("5. Análisis landscape (Post - Algoritmo de Búsqueda)")
        print("9. Salir")

        opcion = input("Selecciona una opción: \n")

        def mostrar_soluciones(problems_dict,opcion):
            
            if opcion == "4": 
                # Imprime los problemas-solución que se han creado en problems_dict                          
                print("Selecciona una opción:")
                for i, (clave, descripcion) in enumerate(problems_dict.items(), start=1):
                    print(f"{i}. {clave}")
            
            if opcion == "5": 
                # Imprime los problemas-solución que se han creado en problems_dict
                # pero que fueron construidos por algoritmos de búsqueda
                print("Estas son las soluciones construidas por algoritmos de búsqueda.")
                print("Selecciona una opción:")
                for i, (clave, descripcion) in enumerate(problems_dict.items(), start=1):
                    if "Nulo" != descripcion.objective:
                        if ('Exacta' != descripcion.tecnica
                            and 'Aproximación' != descripcion.tecnica):                    
                            print(f"{i}. {clave}")
         
        
        def mostrar_operadores():
            while True:
                # Solicito al usuario la función objetivo que desea utilizar
                _menu_options = {
                '1': 'Incremento1_decremento1_exhaust', 
                '2': 'Incremento1_exhaust', 
                '3': 'Incremento1_all', 
                '4': 'Incremento2_decremento1_exhaust', 
                '5': 'Incremento2_decremento2_exhaust',  
                '6': 'Incremento3_decremento3_exhaust',  
                '7': 'Chain_reaction_exhaust_plus_minus', 
                '8': 'Chain_reaction_exhaust_minus_plus', 
                '10': 'Salir al menú anterior'
                }   
            
                print("\n----------------------------------------------------------")
                print("Menú operadores de permutación")
                print("mostrar_operadores@solutions.py")
                print("----------------------------------------------------------\n")
                print("Definir operador:")
                for i,j in _menu_options.items():
                    print (f"{i}. {j}")            
                operador = input("Selecciona una opción: \n")
                if operador == '\n':
                    print("Opción no válida. Inténtalo de nuevo.")
                
                elif operador.isdigit():
                    if int(operador) < 10:
                        return str.lower(_menu_options[str(operador)])
                    
                    elif operador == 10:

                        break
                  
                else:
                    print("Opción no válida. Inténtalo de nuevo.")
             

        if opcion == "1":
            print("Has seleccionado la Opción 1.")
            print(textwrap.dedent(""" \
                  ----------------------------------------------------------
                  Vamos a ingresar tu propia solución.
                  ----------------------------------------------------------
                 
                  Se utilizarán las variables sigma_jk, tao_ijk, z_ijk
                  del archivo de datos de excel.
                  
                  Los flujos entre nodos de servicio 
                  se calculan dividiendo el flujo saliente de cada nodo 
                  en partes iguales. Consulta la matriz df_arcos."""))
            create_problem_object(
                network_original, problems_dict, name_problem="solución_subóptima")

            print("\nSe ha cargado tu solución propia.")
            input("Pulsa una tecla para continuar.")

        elif opcion == "2":
            print("Has seleccionado la Opción 2.")
            print(textwrap.dedent(""" \
                  ----------------------------------------------------------
                  Vamos a obtener soluciones mono-objetivo.
                  ----------------------------------------------------------                
                  
                  A continuación ingresarás a un menú para escoger
                  la función objetivo y el solver respectivo.
                  """))

            # Creo el objeto problem
            create_problem_object(
                network_original, problems_dict, name_problem="temporal")
            current_solution = problems_dict["temporal"]

            # Pido al usuario el objetivo
            current_solution.optimizar=True
            current_solution.menu_mono_optimization(new_network=current_solution.network_copy,
                                                     problems_dict=problems_dict,)
            
            # El usuario ha seleccionado optimizar y la técnica 
            
            # Se ha escogido optimización exacta
            if current_solution.optimizar==True and current_solution.tecnica=='Exacta':
                current_solution.exact_solution()    
                problems_dict[current_solution.name_problem]=current_solution
                
            # Se ha escogido optimización Aproximada
            if current_solution.optimizar==True and current_solution.tecnica=="Aproximación":
                #current_solution.initial_solution(network_original)
                #current_solution.fix_initial_solution()
                current_solution = initial_solution.initial_solution(current_solution,network_original)
                current_solution.set_solution_excel()
                current_solution = initial_solution.fix_initial_solution(current_solution)  # Corrige la solución inicial existente para evitar que tenga rho > 1
                current_solution.set_solution_excel()
                problems_dict[current_solution.name_problem]=current_solution
            
            
            # Se ha escogido Local Search
            if current_solution.optimizar==True and current_solution.tecnica=="Local_Search":
                
                print("\nHas seleccionado escogido usar Local Search.")
                print("\n----------------------------------------------------------")

                print ("Estos son los operadores de permutación:\n")
                operador=mostrar_operadores() #Escojo el operador de permutación
                current_solution.local_search_operator=operador # Lo inserto en el objeto
                current_solution = local_search.local_search(current_solution,network_original)
                current_solution.set_solution_excel()
                problems_dict[current_solution.name_problem]=current_solution

            # Se ha escogido Tabu Search
            if current_solution.optimizar==True and current_solution.tecnica=="Tabu_Search":
                
                print("\nHas seleccionado escogido usar Tabu Search.")
                print("\n----------------------------------------------------------")
            
                print ("Estos son los operadores de permutación:\n")
                operador=mostrar_operadores() #Escojo el operador de permutación
                current_solution.local_search_operator=operador # Lo inserto en el objeto
                current_solution = tabu_search.tabu_search(current_solution,network_original)
                current_solution.set_solution_excel()
                problems_dict[current_solution.name_problem]=current_solution


            # Se ha escogido VND
            if current_solution.optimizar==True and current_solution.tecnica=="VND":
                
                print("\nHas seleccionado escogido usar VND.")
                print("\n----------------------------------------------------------")
            
                print ("No hay necesidad de seleccionar operadores de permutación:\n")
                #operador=mostrar_operadores() #Escojo el operador de permutación
                #current_solution.local_search_operator=operador # Lo inserto en el objeto
                current_solution = vnd.vnd_search(current_solution,network_original)
                current_solution.set_solution_excel()
                problems_dict[current_solution.name_problem]=current_solution
                
            
            # Se ha escogido GVNS
            if current_solution.optimizar==True and current_solution.tecnica=="GVNS":
                
                print("\nHas seleccionado escogido usar GVNS.")
                print("\n----------------------------------------------------------")
            
                print ("No hay necesidad de seleccionar operadores de permutación:\n")
                #operador=mostrar_operadores() #Escojo el operador de permutación
                #current_solution.local_search_operator=operador # Lo inserto en el objeto
                current_solution = gvns.gvns_search(current_solution,network_original)
                current_solution.set_solution_excel()
                problems_dict[current_solution.name_problem]=current_solution

            print ("Los resultados de la optimización se guardaron en salida_optimizacion.xlsx.")               
            input("Pulsa una tecla para continuar.")

        elif opcion == "4":
            print("\nHas seleccionado la Opción 4.")
            print("\n----------------------------------------------------------")
            print ("KPI de soluciones y gráficos.")
            print("----------------------------------------------------------")
            print ("Estas son las soluciones construidas:\n")
            
            # Listar las soluciones existentes
            mostrar_soluciones(problems_dict,opcion)

            # Obtener la elección del usuario
            try:
                from hcndp import kpi
                from hcndp import figures
                from hcndp import export
                
                numero_elegido = int(input("Ingresa el número de la solución elegida: "))
                solucion_elegida = list(problems_dict.keys())[numero_elegido - 1]
                
                # Realizar el procedimiento con la opción seleccionada
                print(f"\nRealizando el procedimiento para la opción: {solucion_elegida}")
                current_solution=problems_dict[solucion_elegida]
                
                # Creo carpetas para current solution
                current_solution.create_folders_problem()
                
                if current_solution.objective=="Nulo":
                    # Si no hay función objetivo (Solución ingresada por usuario)
                    _post_optima=False
                    kpi.calculate_kpi(current_solution,_post_optima)
                    print (f"Se calcularon los KPI para la solución {solucion_elegida}.")
                    print (f"\nAhora escoge el gráfico que deseas para la solución {solucion_elegida}")
                    figures.show_menu_figures(current_solution)
                    export.export_data(current_solution.network_copy)
                    export.create_index_sheet(current_solution.network_copy)
                    #print (f"\Gráficos y archivo de datos exportados a la carpeta de la soución {solucion_elegida}")

                else:
                    # Si hay función objetivo (resultado de optimización)
                    # network_copy no se ha modificado. 
                    # Lo lleno con datos de Excel (si Exacta) o de problem(si Aproximacion)
                    # ACtualizo los sigma, los phi y y los pi_ijkjk
                    # Actualizo las matrices de solution.network_copy
                    current_solution.network_copy.tecnica=current_solution.tecnica
                    current_solution.network_copy.merge_niveles_capac(_post_optima=True, current_solution=problems_dict[solucion_elegida])
                    current_solution.network_copy.create_df_asignacion(_post_optima=True, current_solution=problems_dict[solucion_elegida])
                    current_solution.network_copy.create_df_probs_kk()
                    current_solution.network_copy.create_df_arcos(_post_optima=True, current_solution=problems_dict[solucion_elegida])
                    
                    kpi.calculate_kpi(problems_dict[solucion_elegida],_post_optima=True)
                    print (f"Se calcularon los KPI para la solución {solucion_elegida}.")
                    print (f"\Ahora escoge el gráfico que deseas para la solución {solucion_elegida}")
                    figures.show_menu_figures(problems_dict[solucion_elegida])
                    export.export_data(problems_dict[solucion_elegida].network_copy)
                    export.create_index_sheet(problems_dict[solucion_elegida].network_copy)
                    print (f"\Gráficos y archivo de datos exportados a la carpeta de la soución {solucion_elegida}")
            
            except (ValueError, IndexError) as e:
                print (e)
                print("Error: Ingresa un número válido de la lista.")


        elif opcion == "5":
            print("\nHas seleccionado la Opción 5.")
            print("\n----------------------------------------------------------")
            print ("Análisis del Landscape (Post - Algoritmo de búsqueda.")
            print("----------------------------------------------------------")

            # Listar las soluciones existentes
            mostrar_soluciones(problems_dict,opcion)

            # Obtener la elección del usuario
            try:
                from hcndp import kpi
                from hcndp import figures
                from hcndp import export
                
                numero_elegido = int(input("Ingresa el número de la solución elegida: "))
                solucion_elegida = list(problems_dict.keys())[numero_elegido - 1]
                current_solution=problems_dict[solucion_elegida]
                landscape.landscape(current_solution)                
             
            except (ValueError, IndexError) as e:
                 print (e)
                 print("Error: Ingresa un número válido de la lista.")  



        elif opcion == "9":
            print("Saliendo del programa.")
            break

        else:
            print("\nOpción no válida. Inténtalo de nuevo.")

# %% <codecell> Crear objeto solución


def create_problem_object(network_original, problems_dict, name_problem):

    # Creamos un objeto solution
    new_name_problem = name_problem
    problems_dict[new_name_problem] = Problem(name_problem=new_name_problem,
                                                 objective="Nulo",
                                                 name_network=network_original.name)
    problem = problems_dict[new_name_problem]

    # Inserto una copia de network original en la solución como network_copy
    problem.insert_network_object(network_original)
    
    # Actualizo las matrices de solution.network_copy
    problem.network_copy.merge_niveles_capac(_post_optima=False)
    problem.network_copy.create_df_asignacion(_post_optima=False)
    problem.network_copy.create_df_probs_kk()
    problem.network_copy.create_df_arcos(_post_optima=False)
    
    # Actualizar nombre de solution.network_copy
    problem.network_copy.name_problem=problem.name_problem
    #print ("Se ha creado un objeto tipo 'problema' y se cargó en el diccionario de problemas.")
    
    
# %% <codecell> Clase Problem


class Problem:

    def __init__(self, name_problem, objective, name_network):
        self.name_problem = name_problem
        self.objective = objective
        self.name_network = name_network
        self.state=""

# %% Funciones complementarias
    def insert_network_object(self, network_original):
        import copy
        self.network_copy = copy.deepcopy(network_original)

    def menu_mono_optimization(self, new_network, problems_dict):
        while True:
            print("\n----------------------------------------------------------")
            print("Menú de método de búsqueda")
            print("menu_mono_optimization@solutions.py")
            print("----------------------------------------------------------\n")
            print("1. Solución exacta (Optimización)")
            print("2. Solución inicial (Aproximación)")
            print("3. Solución por búsqueda local (Local Search)")
            print("4. Solución por búsqueda tabú (Tabu Search)")
            print("5. Solución por búsqueda con vecindario variable descendente(VND)")
            print("6. Solución por búsqueda con vecindario variable general(BVNS)")
            
            print("10. Regresar al menú anterior")

            opcion1 = input("Selecciona una opción: \n")
            if opcion1 == "1":
                print("Has seleccionado la opción 1.")
                print("Búsqueda por métodos exactos (Optimización).")
                self.optimizar=True
                self.tecnica="Exacta"
                
                # Presento menú de funciones objetivo disponibles
                objective_and_description = new_network.get_objective_function()
                
                if self.network_copy.optimizar==True:
                    self.objective = objective_and_description[0]
                    self.description_objective = objective_and_description[1]
                    self.name_problem = objective_and_description[1]+" "+self.tecnica
    
                    # Actualizo nombre del problema en problems_dict (Ya no es "temporal")
                    clave_temporal = 'temporal'
                    solucion_temporal = problems_dict[clave_temporal]
                    problems_dict[solucion_temporal.name_problem] = problems_dict.pop(
                        clave_temporal)
                    problems_dict[solucion_temporal.name_problem].network_copy.name_problem = \
                        problems_dict[solucion_temporal.name_problem].name_problem
                    
                    print (f"Se ha actualizado el objeto {problems_dict[solucion_temporal.name_problem].name_problem}")
                    break 
                
                elif self.network_copy.optimizar==False:
                    self.optimizar=False
                    break
                objective_and_description = new_network.get_objective_function()
                
            elif opcion1 == "2":
                print("Has seleccionado la Opción 2.")
                print ("Búsqueda de solución inicial.")
                self.optimizar=True
                self.tecnica="Aproximación"
                # Presento menú de funciones objetivo disponibles
                objective_and_description = new_network.get_objective_function()
                
            
            elif opcion1 == "3":
                print("Has seleccionado la Opción 3.")
                print ("Búsqueda local.")
                self.optimizar=True
                self.tecnica="Local_Search"
                # Presento menú de funciones objetivo disponibles
                objective_and_description = new_network.get_objective_function()
                
            elif opcion1 == "4":
                 print("Has seleccionado la Opción 4.")
                 print ("Búsqueda tabú.")
                 self.optimizar=True
                 self.tecnica="Tabu_Search"
                 # Presento menú de funciones objetivo disponibles
                 objective_and_description = new_network.get_objective_function()
            
            elif opcion1 == "5":
                 print("Has seleccionado la Opción 5.")
                 print ("VND.")
                 self.optimizar=True
                 self.tecnica="VND"
                 # Presento menú de funciones objetivo disponibles
                 objective_and_description = new_network.get_objective_function()
            
            elif opcion1 == "6":
                 print("Has seleccionado la Opción 6.")
                 print ("GVNS.")
                 self.optimizar=True
                 self.tecnica="GVNS"
                 # Presento menú de funciones objetivo disponibles
                 objective_and_description = new_network.get_objective_function()
            
            elif opcion1 == "10":
                print("Has seleccionado la Opción 10.")
                self.optimizar=False
                break

            else:
                print("Opción no válida. Inténtalo de nuevo.")

            if self.network_copy.optimizar==True and self.tecnica !="Exacta":
                 self.objective = objective_and_description[0]
                 self.description_objective = objective_and_description[1]
                 self.name_problem = objective_and_description[1]+" "+self.tecnica
                 
                 # Actualizo nombre de la solución en solution_dict (Ya no es "temporal")
                 clave_temporal = 'temporal'
                 solucion_temporal = problems_dict[clave_temporal]
                 
                 problems_dict[solucion_temporal.name_problem] = problems_dict.pop(
                     clave_temporal)
                 problems_dict[solucion_temporal.name_problem].network_copy.name_problem = \
                     problems_dict[solucion_temporal.name_problem].name_problem
                 print (f"Se ha actualizado el objeto {problems_dict[solucion_temporal.name_problem].name_problem}")
            
                 break 
             
            elif self.network_copy.optimizar==False:
                 self.optimizar=False
                 break
             
            objective_and_description = new_network.get_objective_function()


    def solve_gurobi(self):
        
        network = self.network_copy
        model = self.pyo_model
        instance = self.pyo_model.instance
        #print ("Instancia del modelo de optimización")
        #instance.pprint()
        # Construir el modelo
        #modelo_construido = modelo.create_instance()
        
        # Fijar el valor de la variable para una combinación específica de subíndices
        #valor_fijo = 20
        #instance.sigma.fix(valor_fijo)

        # Crear un objeto StringIO para capturar la salida
        #captura_salida = io.StringIO()
        
        # Guardar la salida estándar original
        #***old_stdout = sys.stdout
        
        # Las líneas siguientes ayudan a acelerar Gurobi. 
        # Fijan valores de variables antes de iniciar la solución.
        #Si s_jk es cero, entonces sigma_sjk es cero
        for j in instance.J:
            for k in instance.K:
                for i in instance.I:
                    if instance.s[j,k]==0:
                        instance.sigma[j,k].fix(0)
                        instance.rho_jk[j,k].fix(0)
                        instance.l_ijk[i,j,k].fix(0)
                        instance.beta_jk[j,k].fix(0)
                        instance.theta[j,k].fix(0)
                    
        #Si w_ij es cero, entonces tao_ijk es cero
        for i in instance.I:
            for j in instance.J:
                for k in instance.K:
                    if instance.w[i,j]==0:
                        instance.tao[i,j,k].fix(0)
                             
        
        # Si x_jj es cero, entonces phi_ijkjk es cero
        for i in instance.I:
            for j in instance.J:
                for k in instance.K:
                    for jp in instance.J:
                        for kp in instance.K:
                            if instance.x[j,jp]==0:
                                instance.fi[i,j,k,jp,kp].fix(0)
                                
        
        # Si r_kk es cero, entonces phi_ijkjk es cero
        for i in instance.I:
            for j in instance.J:
                for k in instance.K:
                    for jp in instance.J:
                        for kp in instance.K:
                            if instance.r_q[k,kp]==0:
                                instance.fi[i,j,k,jp,kp].fix(0)
                                
                                
        
        
        #try:
        # Redirigir la salida estándar al objeto StringIO
        #sys.stdout = captura_salida
    
        # Configurar el solucionador con tee=True      
        # Resolver el modelo con salida detallada
        opt = pyo.SolverFactory('gurobi')
        opt.options['NonConvex'] = 2
        opt.options['TimeLimit'] = 3600
        
        opt.options['MIPFocus'] = 3 #1 Find feasible, 2 prove optimality, 3 very slow
        # opt.options['Heuristics'] = 0.20  #Tiempo usado en heurísticas
        # opt.options['BarConvTol']=1 ### Genera cambbio paulatino
        # opt.options['BarQCPConvTol'] = 1  ### Genera cambbio paulatino
        # opt.options['FeasibilityTol']=0.01 # Primal feasibility tolerance 
        # opt.options['IntFeasTol']=0.1 # Integer feasibility tolerance   ##### Generó cambio importante
        # opt.options['Method']=-1 # Default -1 Algoritmo para modelos continuos 4
        # opt.options['MarkowitzTol']=0.9 # Threshold pivoting tolerance 
        # opt.options['MIQCPMethod']= 1 # 1 outer approx, 0 continuous qcp relax -1 chooses automatically ###################
        # opt.options["NumericFocus"]=0
        # opt.options['OptimalityTol']=0.01 # Dual feasibility tolerance ###### Este generó un cambio importante
        opt.options['MIPGap']=0.01
        # opt.options['SimplexPricing']=-1
        # opt.options['PreQLinearize']=1 # 1 produce a MILP reformulation with strong LP relaxation
        # opt.options['GomoryPasses']= 1 #-1 Default
        # opt.options['PrePasses']= 5 # -1 Default
        # opt.options['Presolve']= 2 #2 agresive #1 conservative Generó cambio importnte con nivel 2
        # opt.options['ScaleFlag']=2 #,1,2,3
        # opt.options['Threads'] = 4  # Usa 4 núcleos
        opt.options['LogFile'] = 'gurobi_log.log'

        global out
        out = 0
        _output = os.getcwd()+'/output/'+network.name+'/'
        
        if self.name_problem == "pareto_front":
            # Crear directorio
            _output = os.getcwd()+'/output/'+'pareto_front_'+str(self.epsilon)+'/'
            if not os.path.exists(_output):
                # Crea el directorio
                os.makedirs(_output)
        
        # Capturar el tiempo antes de resolver
        start_time = time.process_time()

        results = opt.solve(instance, tee=False, warmstart=False)#, report_timing=False)
        #salida_tee = captura_salida.getvalue()
            #,logfile=_output+'logfile_name.log')
        #print (f'Función objetivo alcanzada= {pyo.value(instance.obj)}')   
        
        # Capturar el tiempo después de resolver
        end_time = time.process_time()

        # Calcular el tiempo de CPU utilizado
        cpu_time = end_time - start_time

        # Guardar el tiempo de CPU en una variable
        tiempo_cpu = cpu_time
        
        #print ("Tiempo solución Gurobi:", tiempo_cpu)
               
        # Accessing solver status: http://www.pyomo.org/blog/2015/1/8/accessing-solver
        if (results.solver.status == pyo.SolverStatus.ok) and (results.solver.termination_condition == pyo.TerminationCondition.optimal):
            out = "optimal and feasible"
            # results.write()
            #print(results)
            # instance.display()
            # instance.pprint()
            #instance.rho_max.pprint()
            #instance.alpha_min.pprint()
            #instance.delta_min.pprint()
            model.solution = {}
            model.solution['out'] = "optimal and feasible"
            model.solution['results'] = results
            model.solution['rho_max'] = pyo.value(instance.rho_max)
            model.solution['alpha_min'] = pyo.value(instance.alpha_min)
            model.solution['delta_min'] = pyo.value(instance.delta_min)
            model.solution['Func_obj'] = pyo.value(instance.obj)
            model.solution['Tiempo'] = tiempo_cpu
            
            self.value_optimal_solution=model.solution
            # Export to LP
            #instance.write(_output+"model.lp", format='lp')
                
            
        if results.solver.termination_condition == pyo.TerminationCondition.infeasible:
            out = "infeasible"
            print("infeasible")
        if results.solver.termination_condition == pyo.TerminationCondition.infeasibleOrUnbounded:
            out = "infeasible"
            print("Infeasible or Unbounded")
        if results.solver.termination_condition == pyo.TerminationCondition.maxTimeLimit:
            out = "MaxTimeLimit"
            print("MaxTimeLimit")
        #else:
        #    out = "error"
        print("Solver Status: ",  results.solver.status)
        self.out=out
        
            # Obtener la salida capturada como texto
            #salida_tee = captura_salida.getvalue()
            #self.salida_tee=salida_tee        
        #finally:
        #    # Restaurar la salida estándar
        #    sys.stdout = old_stdout
        return results
        
    def solve_ipopt(self):
        #network = self.network_copy
        model = self.pyo_model
        instance = self.pyo_model.instance

        # Solución por ipopt
        # Instalar ipopt así: conda install -c conda-forge ipopt=3.11.1


        opt = pyo.SolverFactory('ipopt')

        #_output = os.getcwd()+'/output/'+network.name+'/'

        #results=opt.solve(instance, tee=True, logfile=_output+'logfile_name.log')

        results = opt.solve(instance,
                            tee=True
                            )
        #print(results)
        instance.obj.pprint()
        instance.rho_max.pprint()
        instance.alpha_min.pprint()
        instance.delta_min.pprint()
        model.solution = {}
        model.solution['out'] = "optimal and feasible"
        model.solution['results'] = results
        model.solution['rho_max'] = pyo.value(instance.rho_max)
        model.solution['alpha_min'] = pyo.value(instance.alpha_min)
        model.solution['delta_min'] = pyo.value(instance.delta_min)
        
    #def get_degrees_freedom(self):
    #    instance = self.pyo_model.instance
        # Todo: import the degrees_of_freedom function from the idaes.core.util.model_statistics package
        
    #    print("Degrees of Freedom =", degrees_of_freedom(instance))
    #    print("Statistics =", report_statistics(instance))
    
    def set_solution_in_object_local_search(self):
        

        #network = self.network_copy

        # Obtengo los valores de las variables según el modelo de optimización
        # Estas líneas se hacen si el método es exacto.
        instance = self.pyo_model.instance
    
        tao_ijk = np.array([[i, j, k, pyo.value(instance.tao[i, j, k])]
                           for i in instance.I for j in instance.J for k in instance.K])
        tao_ijk = pd.DataFrame(tao_ijk,columns=['nombre_I', 'nombre_J', 'servicio_K', 'tao_ijk'])

        self.df_f_ijk=tao_ijk
        self.df_f_ijk['tao_ijk']=self.df_f_ijk['tao_ijk'].astype(float)
        
        #h_ik = np.array([[i, k, pyo.value(instance.h[i, k])]
        #                for i in instance.I for k in instance.K])
        
        l_ijk = np.array([[i, j, k, pyo.value(instance.l_ijk[i, j, k])]
                         for i in instance.I for j in instance.J for k in instance.K])
        l_ijk  = pd.DataFrame(l_ijk,columns=['nombre_I', 'nombre_J', 'servicio_K', 'lambda_ijk'])

        self.df_asignacion=l_ijk
        self.df_asignacion['lambda_ijk']=self.df_asignacion['lambda_ijk'].astype(float)

        
        #c_jk = np.array([[j, k, pyo.value(instance.c[j, k])]
        #                for j in instance.J for k in instance.K])
        #s_jk = np.array([[j, k, pyo.value(instance.s[j, k])]
        #                for j in instance.J for k in instance.K])

        fi_ijkjk = np.array([[i, j, k, jp, kp, pyo.value(instance.fi[i, j, k, jp, kp])]
                            for i in instance.I for j in instance.J for k in instance.K for jp in instance.J for kp in instance.K])
        fi_ijkjk   = pd.DataFrame(fi_ijkjk,columns=['nombre_I', 
                                                    'nombre_J', 'servicio_K', 
                                                    'nombre_Jp', 'servicio_Kp', 
                                                    'ϕ'])
        self.df_fi_ijkjk=fi_ijkjk
        self.df_fi_ijkjk['ϕ']=self.df_fi_ijkjk['ϕ'].astype(float)

        df_l_ijk = pd.DataFrame(
            l_ijk, columns=['nombre_I', 'nombre_J', 'servicio_K', 'lambda_ijk'])
        fi_jkjk = pd.DataFrame(fi_ijkjk, columns=[
                               'nombre_I', 'nombre_J', 'servicio_K', 'nombre_Jp', 'servicio_Kp', 'ϕ'])
        fi_jkjk=fi_jkjk.rename(columns={'ϕ':'fi_jkjk'})
        fi_jkjk = fi_jkjk.merge(
            df_l_ijk, on=['nombre_I', 'nombre_J', 'servicio_K'], how='left')
        fi_jkjk['fi_jkjk'] = pd.to_numeric(fi_jkjk['fi_jkjk'])
        fi_jkjk['lambda_ijk'] = pd.to_numeric(fi_jkjk['lambda_ijk'])
        fi_jkjk = fi_jkjk.groupby(
            ['nombre_J', 'servicio_K', 'nombre_Jp', 'servicio_Kp']).sum(['lambda_ijk'])

        prob_fi_jkjk = fi_jkjk.groupby(
            ['nombre_J', 'servicio_K', 'nombre_Jp', 'servicio_Kp']).sum(['lambda_ijk'])
        prob_fi_jkjk['fi_jkjk'] = prob_fi_jkjk['fi_jkjk'] / \
            prob_fi_jkjk['lambda_ijk']
        prob_fi_jkjk.fillna(0, inplace=True)
        prob_fi_jkjk.reset_index(inplace=True)

        #fi_jkjk = pd.DataFrame(fi_ijkjk,columns=['nombre_I','nombre_J','servicio_K','nombre_Jp','servicio_Kp','fi_jkjk'])
        # fi_jkjk['fi_jkjk']=pd.to_numeric(fi_jkjk['fi_jkjk'])
        # fi_jkjk=fi_jkjk.groupby(['nombre_J','servicio_K','nombre_Jp','servicio_Kp']).sum().reset_index()
        #fi_jkjk['fi_jkjk']=fi_jkjk.groupby(['nombre_J','servicio_K']).apply(lambda x: x['fi_jkjk']/x['fi_jkjk'].sum()).reset_index()['fi_jkjk']
        prob_fi_jkjk = prob_fi_jkjk.to_numpy()
        prob_fi_jkjk = prob_fi_jkjk[:, :-1]
        prob_fi_jkjk= pd.DataFrame(prob_fi_jkjk,columns=['nombre_J', 'servicio_K', 
                                                    'nombre_Jp', 'servicio_Kp', 
                                                    'Probs'])

        self.df_prob_fi_jkjk=prob_fi_jkjk
        self.df_prob_fi_jkjk['Probs']=self.df_prob_fi_jkjk['Probs'].astype(float)

        # alpha_ik = np.array([[i, k, pyo.value(instance.alpha_ik[i, k])]
        #                     for i in instance.I for k in instance.K])
        # rho_jk = np.array([[j, k, pyo.value(instance.rho_jk[j, k])]
        #                   for j in instance.J for k in instance.K])
        # f_ijk = np.array([[i, j, k, pyo.value(instance.tao[i, j, k])]
        #                  for i in instance.I for j in instance.J for k in instance.K])
        l_jk = np.array([[j, k, sum(pyo.value(instance.l_ijk[i, j, k])
                        for i in instance.I)] for j in instance.J for k in instance.K])
        l_jk  = pd.DataFrame(l_jk,columns=['nombre_J', 'servicio_K', 'lambda_jk'])

        self.df_l_jk=l_jk
        
        sigma_jk = np.array([[j, k, pyo.value(instance.sigma[j, k])]
                            for j in instance.J for k in instance.K])
        sigma_jk= pd.DataFrame(sigma_jk,columns=['nombre_J', 'servicio_K', 'sigma_jk'])
        

        self.df_sigma=sigma_jk
        self.df_sigma['sigma_jk']=self.df_sigma['sigma_jk'].astype(float)

        # theta_jk = np.array([[j, k, pyo.value(instance.theta[j, k])]
        #                     for j in instance.J for k in instance.K])
    
        # Borro el dataframe df_prob_fi_ijkjk porque no logro actualizarlo
        try: 
            del self.df_prob_fi_ijkjk
        except AttributeError as e:
            pass
    
    def set_solution_in_object_network_repr(self):
        # Actualizo los valores de lambda en las matrices en cada nodo oferta
        for _,_j in self.network_repr.nodes_supply.items():
            if _j.service!='k00':
                _j.matriz_λ=self.df_asignacion.set_index(['nombre_I','nombre_J','servicio_K']).loc[(slice(None),_j.place,_j.service),:]
                _j.matriz_λ.reset_index(inplace=True)
                _j.matriz_λ = _j.matriz_λ.rename(columns={'lambda_ijk': 'λ_ijk'})
                
        
    
    def set_solution_excel(self):

        # network = self.network_copy

        # Obtengo los valores de las variables según el modelo de optimización
        # Estas líneas se hacen si el método es exacto.
        if self.tecnica=='Exacta':
            instance = self.pyo_model.instance
        
            tao_ijk = np.array([[i, j, k, pyo.value(instance.tao[i, j, k])]
                               for i in instance.I for j in instance.J for k in instance.K])
            h_ik = np.array([[i, k, pyo.value(instance.h[i, k])]
                            for i in instance.I for k in instance.K])
            l_ijk = np.array([[i, j, k, pyo.value(instance.l_ijk[i, j, k])]
                             for i in instance.I for j in instance.J for k in instance.K])
            c_jk = np.array([[j, k, pyo.value(instance.c[j, k])]
                            for j in instance.J for k in instance.K])
            s_jk = np.array([[j, k, pyo.value(instance.s[j, k])]
                            for j in instance.J for k in instance.K])
    
            fi_ijkjk = np.array([[i, j, k, jp, kp, pyo.value(instance.fi[i, j, k, jp, kp])]
                                for i in instance.I for j in instance.J for k in instance.K for jp in instance.J for kp in instance.K])
            df_l_ijk = pd.DataFrame(
                l_ijk, columns=['nombre_I', 'nombre_J', 'servicio_K', 'lambda_ijk'])
            fi_jkjk = pd.DataFrame(fi_ijkjk, columns=[
                                   'nombre_I', 'nombre_J', 'servicio_K', 'nombre_Jp', 'servicio_Kp', 'fi_jkjk'])
            fi_jkjk = fi_jkjk.merge(
                df_l_ijk, on=['nombre_I', 'nombre_J', 'servicio_K'], how='left')
            fi_jkjk['fi_jkjk'] = pd.to_numeric(fi_jkjk['fi_jkjk'])
            fi_jkjk['lambda_ijk'] = pd.to_numeric(fi_jkjk['lambda_ijk'])
            fi_jkjk = fi_jkjk.groupby(
                ['nombre_J', 'servicio_K', 'nombre_Jp', 'servicio_Kp']).sum(['lambda_ijk'])
    
            prob_fi_jkjk = fi_jkjk.groupby(
                ['nombre_J', 'servicio_K', 'nombre_Jp', 'servicio_Kp']).sum(['lambda_ijk'])
            prob_fi_jkjk['fi_jkjk'] = prob_fi_jkjk['fi_jkjk'] / \
                prob_fi_jkjk['lambda_ijk']
            prob_fi_jkjk.fillna(0, inplace=True)
            prob_fi_jkjk.reset_index(inplace=True)
    
            #fi_jkjk = pd.DataFrame(fi_ijkjk,columns=['nombre_I','nombre_J','servicio_K','nombre_Jp','servicio_Kp','fi_jkjk'])
            # fi_jkjk['fi_jkjk']=pd.to_numeric(fi_jkjk['fi_jkjk'])
            # fi_jkjk=fi_jkjk.groupby(['nombre_J','servicio_K','nombre_Jp','servicio_Kp']).sum().reset_index()
            #fi_jkjk['fi_jkjk']=fi_jkjk.groupby(['nombre_J','servicio_K']).apply(lambda x: x['fi_jkjk']/x['fi_jkjk'].sum()).reset_index()['fi_jkjk']
            # fi_jkjk=fi_jkjk.fillna(0)
            prob_fi_jkjk = prob_fi_jkjk.to_numpy()
            prob_fi_jkjk = prob_fi_jkjk[:, :-1]
    
            alpha_ik = np.array([[i, k, pyo.value(instance.alpha_ik[i, k])]
                                for i in instance.I for k in instance.K])
            rho_jk = np.array([[j, k, pyo.value(instance.rho_jk[j, k])]
                              for j in instance.J for k in instance.K])
            f_ijk = np.array([[i, j, k, pyo.value(instance.tao[i, j, k])]
                             for i in instance.I for j in instance.J for k in instance.K])
            l_jk = np.array([[j, k, sum(pyo.value(instance.l_ijk[i, j, k])
                            for i in instance.I)] for j in instance.J for k in instance.K])
            sigma_jk = np.array([[j, k, pyo.value(instance.sigma[j, k])]
                                for j in instance.J for k in instance.K])
            theta_jk = np.array([[j, k, pyo.value(instance.theta[j, k])]
                                for j in instance.J for k in instance.K])
        
        # Obtengo los datos si la solución fue por aproximación
        if self.tecnica != 'Exacta':
        #if self.tecnica=='Aproximación' or self.tecnica=="Local_Search":
            l_ijk=self.df_asignacion
            l_ijk.columns = ['0', '1', '2','3']
            
            l_jk = l_ijk.set_index(['1','2']).groupby(level=['1', '2']).sum()
            l_jk=l_jk.drop(columns=['0']).reset_index()
            #l_jk = self.df_l_jk.reset_index()
            l_jk.columns = ['0', '1', '2']
                    
            f_ijk=self.df_f_ijk
            f_ijk.columns = ['nombre_I','nombre_J','servicio_K','tao_ijk']
            
            prob_fi_jkjk=self.df_prob_fi_jkjk.reset_index()
            try:
                prob_fi_jkjk = prob_fi_jkjk[['nombre_J','servicio_K','nombre_Jp','servicio_Kp','π_jkjk']]
            except Exception:
                pass
            
            if 'index' in prob_fi_jkjk.columns:
                prob_fi_jkjk.drop(columns=['index'], inplace=True)

            prob_fi_jkjk.columns = ['nombre_J','servicio_K','nombre_Jp','servicio_Kp','Probs']
            prob_fi_jkjk = prob_fi_jkjk[prob_fi_jkjk['servicio_K'] != 'k00']
            
            sigma_jk=self.df_sigma
            sigma_jk.columns = ['0','1','2']
            
            fi_ijkjk=self.df_fi_ijkjk
            fi_ijkjk = fi_ijkjk[['nombre_I','nombre_J','servicio_K','nombre_Jp','servicio_Kp','ϕ']]
            fi_ijkjk.columns = ['nombre_I','nombre_J','servicio_K','nombre_Jp','servicio_Kp','fi_ijkjk']
            fi_ijkjk=fi_ijkjk[fi_ijkjk['servicio_K'] != 'k00']
                       
        # Escritura de archivo en Excel
        #output = os.getcwd()+'/output/'+network.name+'/salida_optimizacion.xlsx'
        if self.name_problem != "pareto_front":
            output = os.getcwd()+'/output/'+self.name_problem+'/salida_optimizacion.xlsx'
            
            # Verificar si el directorio existe, y si no, crearlo
            directorio = os.path.dirname(output)
            if not os.path.exists(directorio):
                os.makedirs(directorio)
    
            workbook = xlsxwriter.Workbook(output)
            workbook.close()
            path = output
        elif self.name_problem == "pareto_front":            
            output = os.getcwd()+'/output/'+'pareto_front_'+str(self.epsilon)+'/salida_optimizacion.xlsx'
            
            # Verificar si el directorio existe, y si no, crearlo
            directorio = os.path.dirname(output)
            if not os.path.exists(directorio):
                os.makedirs(directorio)
    
            workbook = xlsxwriter.Workbook(output)
            workbook.close()
            path = output
        
        if self.tecnica=='Exacta':
            with pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                writer.workbook = openpyxl.load_workbook(path)
                pd.DataFrame(l_jk).to_excel(writer, sheet_name='l_jk')
                pd.DataFrame(l_ijk).to_excel(writer, sheet_name='l_ijk')
                pd.DataFrame(f_ijk, columns=['nombre_I', 'nombre_J', 'servicio_K', 'tao_ijk']).to_excel(
                    writer, sheet_name='f_ijk')  # Corresponde a los tao_ijk
    
                pd.DataFrame(prob_fi_jkjk, columns=['nombre_J', 'servicio_K', 'nombre_Jp', 'servicio_Kp', 'Probs']).to_excel(
                    writer, sheet_name='prob_fi_jkjk')
    
                pd.DataFrame(sigma_jk).to_excel(writer, sheet_name='sigma')
                pd.DataFrame(fi_ijkjk, columns=['nombre_I', 'nombre_J', 'servicio_K', 'nombre_Jp',
                             'servicio_Kp', 'fi_ijkjk']).to_excel(writer, sheet_name='fi_ijkjk')
                
                # Crear una nueva hoja para el tiempo
                tiempo_data = pd.DataFrame({'Tiempo': [self.pyo_model.solution['Tiempo']]})
                tiempo_data.to_excel(writer, sheet_name='Tiempo')
       
        if self.tecnica != "Exacta":
            with pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                writer.workbook = openpyxl.load_workbook(path)
                l_jk.to_excel(writer, sheet_name='l_jk')
                l_ijk.to_excel(writer, sheet_name='l_ijk')
                f_ijk.to_excel(
                    writer, sheet_name='f_ijk')  # Corresponde a los tao_ijk
    
                prob_fi_jkjk.to_excel(
                    writer, sheet_name='prob_fi_jkjk')
    
                sigma_jk.to_excel(writer, sheet_name='sigma')
                fi_ijkjk.to_excel(writer, sheet_name='fi_ijkjk')
        
        print(f"Se exportó exitosamente el archivo de Excel en {path}.\n")
        
        if self.tecnica == 'Exacta':
            print(f"Explora el objeto {self.name_problem}, y la matriz detailed_solution para mayor detalle de la solución.\n")
            self.detailed_solution={'tao_ijk':tao_ijk,
                                    'h_ik':h_ik,
                                    'l_ijk':l_ijk,
                                    'c_jk':c_jk,
                                    's_jk':s_jk,
                                    'fi_ijkjk':fi_ijkjk,
                                    'df_l_ijk':df_l_ijk,
                                    'fi_jkjk':fi_jkjk,
                                    'prob_fi_jkjk':prob_fi_jkjk,
                                    'alpha_ik':alpha_ik,
                                    'rho_jk':rho_jk,
                                    'f_ijk':f_ijk,
                                    'l_jk':l_jk,
                                    'sigma_jk':sigma_jk,
                                    'theta_jk':theta_jk,
                                    'tiempo':tiempo_data
                                }
    def set_solution_txt(self):

        ##########################################
        #Imprimir resultados en un achivo ####
        ##########################################

        # network = self.network_copy
        instance = self.pyo_model.instance

        #output = os.getcwd()+'/output/'+network.name+'/modeloysolucion.txt'
        output = os.getcwd()+'/output/'+self.name_problem+'/modeloysolucion.txt'
        
        # Verificar si el directorio existe, y si no, crearlo
        directorio = os.path.dirname(output)
        if not os.path.exists(directorio):
            os.makedirs(directorio)

        with open(output, 'w') as output_file:
            # output_file.write(instance.pprint())
            output_file.write("Tiempo de ejecución: "+str(self.pyo_model.solution['Tiempo']))
            output_file.write("\n"+"#"*60+"\n")
            instance.pprint(output_file)
        
        # Abrir el archivo en modo de agregación para agregar 'self.salida_tee'
        #with open(output, 'a') as output_file:
            #output_file.write("\n"+"#"*60)  # Agregar un salto de línea para separación
            #output_file.write(self.salida_tee)  # Agregar la salida tee de Gurobi
            #output_file.write("\n"+str(self.pyo_model.solution['Tiempo']))
            
        output = os.getcwd()+'/output/'+self.name_problem+'/solucion.txt'
        if self.name_problem == "pareto_front":
            output = os.getcwd()+'/output/'+'pareto_front_'+str(self.epsilon)+'/solucion.txt'
  
        # Verificar si el directorio existe, y si no, crearlo
        directorio = os.path.dirname(output)
        if not os.path.exists(directorio):
            os.makedirs(directorio)
        instance.display(output)

        print(f"Se exportó exitosamente el archivo de texto en {output}")

    
    
    def calculate_kpi_before_optim(self,network_copy,_post_optima=False):
        #print ("Calculo los KPI necesarios para la optimización.")
        kpi.set_lambda_jk(self,network_copy,_post_optima=False)
        kpi.set_lambda_ijk(self,network_copy,_post_optima=False)
        kpi.set_phi_ijkjk(self,network_copy)
        kpi.set_prop_tao(self,network_copy)
        kpi.set_prob_k(self,network_copy)
    


# %% Tipo de solución
    def exact_solution(self):
        # Se ha escogido optimización exacta
        if self.optimizar==True and self.tecnica=='Exacta':
            # Calculo la solución óptima
            self.construct_model()
            # Creo el archivo data_dat (network)
            self.network_copy.create_data_dat()
            self.construct_instance()
            self.execute_solver()

    def initial_solution(self,network_original):
        #Creo la representación de la red
       
        network_repr=network.Network_representation(network_original.I,
                                                    network_original.J,
                                                    network_original.K,
                                                    network_original.archivo,
                                                    network_original.file)
        path_repr=network.Path_representation(network_original.K, 
                                              network_original.archivo,
                                              network_original.file)
        
        # Asignación de recurso σ_k para cada nodo de oferta j 
        print ("Inicio la asignación de recursos sigma")
        for _k in path_repr.nodes_services.keys():
            if _k !=  'k00':
                #print ("Asignación de recursos para: ", _k)
                network_repr.asignacion_recursos(path_repr,_k)
                
    
        # Construyo el df_sigma con los resultados de la asignacion
        # Crear listas vacías para almacenar los datos
        places = []
        services = []
        capac_instal_sigma = []

        # Iterar a través del diccionario nodes_supply y extraer los datos
        for node in network_repr.nodes_supply.values():
            places.append(node.place)
            services.append(node.service)
            capac_instal_sigma.append(node.capac_instal_sigma)

        # Crear el DataFrame
        network_repr.df_sigma = pd.DataFrame({
            'place': places,
            'service': services,
            'capac_instal_sigma': capac_instal_sigma
             })            
        network_repr.df_sigma = network_repr.df_sigma.dropna(subset=['capac_instal_sigma'])
        self.df_sigma=network_repr.df_sigma
        self.df_sigma = self.df_sigma.rename(columns={'place':'nombre_J', 'service':'servicio_K', 'capac_instal_sigma':'sigma_jk'})


        # Para k=0 creo valores de λ_ijk0 
        # El valor de λ_ijk0 es igual a la demanda que hay en cada nodo ik0.
        for _,_j in network_repr.nodes_supply.items():
            for _p,_i in _j.matriz_λ.iterrows():
                if _i['nombre_I'].replace('i','j')==_i['nombre_J'] and _i['servicio_K']=='k00':
                    _j.matriz_λ.loc[_p, 'λ_ijk'] = network_repr.nodes_demand[_i['nombre_I']].demand
        # Construyo primera solución 
        # Ordeno el diccionario de ser_ser_R
        path_repr.edges_ser_ser_R=dict(sorted(path_repr.edges_ser_ser_R.items()))
        
        print ("Asignación de flujos ijkkp")
        for _i in path_repr.edges_ser_ser_R.values():
            k=_i.source
            kp=_i.target
        
                # Obtengo los delta ijkk

            network_repr.asignacion_flujos_δ(network_repr,path_repr,k,kp) 
            
                # Solución entre k y kp* obtengo los phi ijkjk
            network_repr.solucion_entre_k_kp(network=network_repr,_k=k,
                                                 _kp=kp,
                                                 archivo=network_original.file)
                
                # Obtengo las aproximaciones de lambda
            network_repr.construyo_λ(network_repr,kp)
            
        # Porcentajes de flujo. Obtengo los pi jk jk
        network_repr.obtencion_π(network_repr)
                    
        # Construyo una matriz g con los arribos externos, es decir los ϕi.j.k0jk
    
        _lista_i=indices("i",network_original.I)
        _lista_j=indices("j",network_original.J)
        _lista_k=indices("k",network_original.K)
        
        _lista = list(product(_lista_i, _lista_j,_lista_k))
        _g=pd.DataFrame(_lista, columns=['nombre_I', 'nombre_J','servicio_K'])
        _g['tao_ijk'] = 0.0
        _g=_g.sort_values(by=['nombre_I', 'nombre_J','servicio_K'])
        self.df_f_ijk=_g
        
        def arribos_externos(poblacion_origen,servicio_origen,nodo_destino):
            suma_acumulativa=0
            for clave,valor in network_repr.edges_sup_sup_X.items():
                if valor.node_demand_pop == poblacion_origen and valor.service_source==servicio_origen and valor.target==nodo_destino :
                    suma_acumulativa += valor.flow_sup_sup_phi
            return suma_acumulativa
            
        _g['tao_ijk']=_g.apply(lambda row: arribos_externos(row['nombre_I'],'k00',row['nombre_J']+row['servicio_K']),axis=1)
        _g=np.array(_g['tao_ijk']) # Lista de arribos externos ijk
        _g=np.reshape(_g,([network_original.I,network_original.J*(network_original.K)])) # Matriz de arribos externos de i por (jk)
        
        
        # Construyo las matrices con las probabilidades π  
        _lista = list(product(_lista_i, _lista_j,_lista_k,_lista_j,_lista_k))
        _π=pd.DataFrame(_lista, columns=['nombre_I', 'nombre_J','servicio_K', 'nombre_Jp','servicio_Kp'])
        _π['π_ijkjk'] = 0.0   
        _π=_π.sort_values(by=['nombre_I','nombre_J', 'servicio_K','nombre_Jp','servicio_Kp'])
        _π['p*π'] = 0.0
        
        for _,_fila in _π.iterrows():
            _i=_fila['nombre_I']
            _j=_fila['nombre_J']
            _k=_fila['servicio_K']
            _jp=_fila['nombre_Jp']
            _kp=_fila['servicio_Kp']
            for _nombre,_arco in network_repr.edges_sup_sup_X.items():
                if  _nombre == _i+_j+_k+_jp+_kp:
                    _π.loc[_,'π_ijkjk'] = _arco.flow_sup_sup_perc
            
            for _nombre,_arco in path_repr.edges_ser_ser_R.items():
                if  _nombre == _k+_kp:
                    _π.loc[_,'p*π'] = _arco.transfer_percentage  * _π.loc[_,'π_ijkjk']           
       
        # Calculo los flujos entrantes lambda ijk basado en redes de Jackson
        _lista = list(product(_lista_i, _lista_j,_lista_k))
        _df_asignacion=pd.DataFrame(_lista, columns=['nombre_I', 'nombre_J','servicio_K'])
        _df_asignacion["lambda_ijk"] = 0.0
        _df_asignacion=_df_asignacion.sort_values(by=['nombre_I','nombre_J', 'servicio_K'])
        _df_asignacion.set_index(['nombre_I', 'nombre_J','servicio_K'],inplace=True)
        
        
        
        #Para cada i calculo el lambda ijk usando Jackson
        _fila=0
        for i in _lista_i:
            probs=_π.loc[(_π['nombre_I']==_lista_i[_fila])]
            probs=probs.sort_values(by=['nombre_J', 'servicio_K','nombre_Jp','servicio_Kp'])
            probs=np.array(probs['π_ijkjk'])
            probs=np.reshape(probs,([network_original.J*(network_original.K),network_original.J*(network_original.K)]))
            _df_asignacion.loc[i,'lambda_ijk']=np.matmul(_g[_fila],np.linalg.inv(np.identity(len(probs))-(probs)))
            _fila+=1
        # Actualizo los valoes de lambda en las matrices en cada nodo oferta
        for _,_j in network_repr.nodes_supply.items():
            if _j.service!='k00':
                _j.matriz_λ=_df_asignacion.loc[(slice(None),_j.place,_j.service),:]
                _j.matriz_λ.reset_index(inplace=True)
                _j.matriz_λ = _j.matriz_λ.rename(columns={'lambda_ijk': 'λ_ijk'})
        
        # Actualizo los delta con base en los nuevos lambda
        print ("Actualización de flujos ijkkp")
        for _i in path_repr.edges_ser_ser_R.values():
            k=_i.source
            kp=_i.target
        
            # Obtengo los delta ijkk
            
            network_repr.asignacion_flujos_δ(network_repr,path_repr,k,kp) 
            
        # Actualizo los phi con base en los nuevos lambda
        network_repr.solucion_flujos_phi_post_Jackson(network_repr)
                
               
        # Actualizo el df_asignacion en network
        
        self.df_asignacion=_df_asignacion.reset_index()
        network_repr.df_asignacion=_df_asignacion
        self.df_l_jk = network_repr.df_asignacion.groupby(level=['nombre_J', 'servicio_K']).sum()
        self.df_l_jk = self.df_l_jk.rename(columns={'lambda_ijk': 'lambda_jk'})
        self.df_l_jk.reset_index(inplace=True)
        
        # Llevar solución a un df_solucion
        _lista=[] 
        for _i,_j in network_repr.edges_sup_sup_X.items():
            _lista.append([_j.node_demand_pop,
                           _j.source,_j.target,
                           _j.flow_sup_sup_perc_ijkjk,
                           _j.flow_sup_sup_phi,
                           _j.flow_sup_sup_perc_jkjk])
        df_solucion = pd.DataFrame(_lista, columns=['nombre_I', 'origen', 'destino','π_ijkjk','ϕ','π_jkjk'])
        
        self.solution=df_solucion
        
        # Preparo los df_prob_fi_jkjk y df_fi_ijkjk
        df_solucion['nombre_J'] = df_solucion['origen'].str[:3]
        df_solucion['servicio_K'] = df_solucion['origen'].str[3:]
        df_solucion['nombre_Jp'] = df_solucion['destino'].str[:3]
        df_solucion['servicio_Kp'] = df_solucion['destino'].str[3:]

        # Eliminar la columna 'origen'
        df_solucion.drop(columns=['origen','destino'], inplace=True)
        
        #Como df_solucion no tiene todas las combinaciones de ijkj'k', tengo que completar la matriz
        _lista_I=indices("i",network_repr.I)
        _lista_J=indices("j",network_repr.J)
        _lista_K=indices("k",network_repr.K)
        _lista_completa = np.array([[i, j, k, jp, kp, 0,0,0]
                            for i in _lista_I for j in _lista_J for k in _lista_K for jp in _lista_J for kp in _lista_K])
        _lista_completa = pd.DataFrame(_lista_completa, 
                                       columns=['nombre_I', 
                                                'nombre_J', 
                                                'servicio_K',
                                                'nombre_Jp',
                                                'servicio_Kp',
                                                'π_ijkjk','ϕ','π_jkjk'])
        _lista_completa['π_jkjk'] = _lista_completa['π_jkjk'].astype(float)
        _lista_completa['ϕ'] = _lista_completa['ϕ'].astype(float)
        _lista_completa['π_ijkjk'] = _lista_completa['π_ijkjk'].astype(float)
        
        #_lista_completa.update(df_solucion)
        df_solucion= pd.merge(_lista_completa, df_solucion, on=['nombre_I', 'nombre_J', 'servicio_K', 'nombre_Jp', 'servicio_Kp'],how='left')
        df_solucion['π_ijkjk'] = df_solucion['π_ijkjk_x'] + df_solucion['π_ijkjk_y']
        df_solucion.drop(['π_ijkjk_x', 'π_ijkjk_y'], axis=1, inplace=True)
        df_solucion['π_jkjk'] = df_solucion['π_jkjk_x'] + df_solucion['π_jkjk_y']
        df_solucion.drop(['π_jkjk_x', 'π_jkjk_y'], axis=1, inplace=True)
        df_solucion['ϕ'] = df_solucion['ϕ_x'] + df_solucion['ϕ_y']
        df_solucion.drop(['ϕ_x', 'ϕ_y'], axis=1, inplace=True)
        df_solucion.fillna(0, inplace=True)
        
        df_prob_fi_ijkjk = copy.deepcopy(df_solucion)
        df_prob_fi_ijkjk.drop(columns=['ϕ','π_jkjk'],inplace=True)
        
        df_fi_ijkjk = copy.deepcopy(df_solucion)
        df_fi_ijkjk.drop(columns=['π_ijkjk','π_jkjk'],inplace=True)
        
        df_prob_fi_jkjk = copy.deepcopy(df_solucion)
        df_prob_fi_jkjk.drop(columns=['ϕ','π_ijkjk'],inplace=True)
        df_prob_fi_jkjk = df_prob_fi_jkjk.groupby(
            ['nombre_J', 'servicio_K', 'nombre_Jp', 'servicio_Kp']).mean(['π_jkjk'])
        
        df_prob_fi_jkjk=df_prob_fi_jkjk.rename(columns={'π_jkjk':'Probs'})
        
        self.df_prob_fi_ijkjk =df_prob_fi_ijkjk 
        self.df_fi_ijkjk=df_fi_ijkjk
        self.df_prob_fi_jkjk =df_prob_fi_jkjk 
        
        # Calculo los rho para cada nodo jk y lo almaceno en nodes_supply
        for _i,_j in network_repr.nodes_supply.items():
            if _j.service != 'k00':
                if _j.matriz_λ['λ_ijk'].sum() == 0:
                    _j.rho=0
                else: 
                    _j.rho = _j.matriz_λ['λ_ijk'].sum()/(_j.capac_instal_sigma*_j.rate)

        self.state="Solucionado Solución Inicial"
        self.network_repr=network_repr
        


        return self
    
    
    def fix_initial_solution(self):
        # Corrige la solución inicial existente para evitar que tenga rho > 1
        
        local_search.calcular_kpi_local_search(self)
        
        while self.network_copy.file['df_capac']['rho'].max() > 1: # Mientras que haya algún nodo sobresaturado
    

            for _i,_j in self.network_repr.nodes_supply.items():
                if _j.service != 'k00' and _j.rho > 1: # Identifico elementos con rho > 1
                    #print (_i,_j.rho)
                    nodo_mayor_congest= _i
                    servicio_congest=_j.service # Identifico el servicio del elemento congestionado
                    menor_congestion=2
                    # Identifico el sigma menos congestionado de servicio_congest
                    for _m,_k in  self.network_repr.nodes_supply.items():
                        if _k.service == servicio_congest and _k.rho < menor_congestion:
                            menor_congestion=_k.rho
                            nodo_menos_congest=_m
                            #print (_m,_k,menor_congestion)
                    # Ajusto sigmas entre nodo_mayor_congest y nodo_menor_congest
                    self.network_repr.nodes_supply[nodo_mayor_congest].capac_instal_disponible -= 1 
                    self.network_repr.nodes_supply[nodo_mayor_congest].capac_instal_sigma += 1
                    
                    self.network_repr.nodes_supply[nodo_menos_congest].capac_instal_disponible += 1 
                    self.network_repr.nodes_supply[nodo_menos_congest].capac_instal_sigma -= 1
                    
                    # Actualizo los nuevos valores de sigma de network_repr en network_copy > file > df_capac
                    for _i,_j in self.network_repr.nodes_supply.items():
                        a = _j.place
                        b = _j.service
                        c = _j.capac_instal_sigma
                    
                        fila = self.network_copy.file['df_capac'].query('nombre_J == @a and servicio_K == @b')
                        self.network_copy.file['df_capac'].loc[fila.index, 'sigma_jk'] = c
                        fila = self.df_sigma.query('nombre_J == @a and servicio_K == @b')
                        self.df_sigma.loc[fila.index, 'sigma_jk'] = c
                        
                    # Calculo los rho para cada nodo jk y lo almaceno en nodes_supply
                    for _i,_j in self.network_repr.nodes_supply.items():
                         if _j.service != 'k00':                                    
                            if _j.matriz_λ['λ_ijk'].sum() == 0:
                                _j.rho=0
                            else: 
                                _j.rho = _j.matriz_λ['λ_ijk'].sum()/(_j.capac_instal_sigma*_j.rate)                        
                    
                    
                    
                    print ("")
                    local_search.calcular_kpi_local_search(self)
                    
        return  self
        # %% Exportar resultados 
        #self.set_solution_excel()
        
    # %% Crear modelo abstracto
    def construct_model(self):

        _menu_options = {
            '1': 'Minimizar congestión máxima (rho)',
            '2': 'Maximizar accesibilidad mínima (alpha)',
            '3': 'Maximizar continuidad mínima (delta)',
            '4': 'Maximizar accesibilidad total (alpha)',
            '5': 'Minimizar usuarios en espera total (Lq_total)',
            '6': 'Maximizar continuidad total (delta total)',
            '7': 'Salir al menú anterior'
        }
        objective = self.objective

        # Creo objeto Model_pyomo dentro de solution
        self.pyo_model = models.Model_pyomo(
            model_abstract=None,
            instance=None,
            data_dat=None,
            solution=None,
            nombre_modelo=self.name_problem)

        # Lleno pyo_model con el modelo abstracto
        self.pyo_model.set_model_abstract(objetivo=objective,
                                          nombre_modelo=self.name_problem,
                                          _menu_options=_menu_options,
                                          new_network=self.network_copy
                                          )

        # Calculo kpi necesarios (solution)
        self.calculate_kpi_before_optim(network_copy=self.network_copy,
                                        _post_optima=False)

        # # Creo el archivo data_dat (network)
        # self.network_copy.create_data_dat()

    # %% Instanciar (model)

    
    def construct_instance(self):
        
        #self.pyo_model.data_dat = os.getcwd()+'/data/'+self.network_copy.name+'/datos.dat'
        self.pyo_model.data_dat = os.getcwd()+'/data/'+self.name_network+'/datos.dat'
        
        self.pyo_model.instance = self.pyo_model.model_abstract.create_instance(self.pyo_model.data_dat)
        
        # Definir expresión de objetivo si es necesario
        if self.objective == 4:
            obj_expr = sum(self.pyo_model.instance.alpha_ik[i, k] * sum(self.pyo_model.instance.h[i, kp]
                          for kp in self.pyo_model.instance.K) for i in self.pyo_model.instance.I
                           for k in self.pyo_model.instance.K) / sum(sum(self.pyo_model.instance.h[i, kp]
                           for kp in self.pyo_model.instance.K) for i in self.pyo_model.instance.I
                           for k in self.pyo_model.instance.K)
            self.pyo_model.instance.obj = pyo.Objective(expr=obj_expr, sense=pyo.maximize)
            
        elif self.objective == 5:
            for j in self.pyo_model.instance.J:
                for k in self.pyo_model.instance.K:
                    self.pyo_model.instance.rho_jk[j, k].setlb(0.000000000000001)
                    self.pyo_model.instance.rho_jk[j, k].value = 0.000000000000001
                    
        elif self.objective == 6:
            obj_expr = sum(self.pyo_model.instance.delta_i[i] * sum(self.pyo_model.instance.h[i, k]
                          for k in self.pyo_model.instance.K) for i in self.pyo_model.instance.I) \
                          / sum(self.pyo_model.instance.h[i, k] for i in self.pyo_model.instance.I
                          for k in self.pyo_model.instance.K)
            self.pyo_model.instance.obj = pyo.Objective(expr=obj_expr, sense=pyo.maximize)


    # %% Resolver
    
    def execute_solver(self):
        if self.tecnica=="Local_Search" or self.tecnica=="Tabu_Search" or self.tecnica=="Aproximación" or\
            self.tecnica=="VND" or self.tecnica=="GVNS":
            if self.objective == 5:
                self.solve_ipopt()
            else:
                self.solve_gurobi()
            #self.get_degrees_freedom()
    
            if self.out != "infeasible":
                # Exportar resultados
                self.set_solution_in_object_local_search()
                self.set_solution_in_object_network_repr()
                #self.network_copy.create_folders()
                #self.set_solution_excel()
                #self.set_solution_txt()
                self.state="Optimizado"
                
            

        else:
            if self.objective == 5:
                self.solve_ipopt()
            else:
                self.solve_gurobi()
                
                print ("Valor función objetivo:",self.value_optimal_solution['Func_obj'])
            #self.get_degrees_freedom()
    
            # Exportar resultados
            self.network_copy.create_folders()
            self.set_solution_excel()
            self.set_solution_txt()
            self.state="Optimizado"

    
    # %% Exportar solución
    def create_folders_problem(self):
        
        if not os.path.exists(os.getcwd()+'/output/'+self.name_problem+'/'):
            # Crea el directorio
            os.makedirs(os.getcwd()+'/output/'+self.name_problem+'/')
            #print(f"Directorio /output/'{self.name_problem}' creado con éxito.")
        else:
            #print(f"El directorio /output/'{self.name_problem}' ya existe.")
            pass
        


def update_solution_post_optima(original_network, new_network):

    print("Ejecutando update_solution_post_optima")
    input()
    new_network.create_folders()
    path = os.getcwd()+'/data/'+original_network.name+'/'+new_network.archivo


    # Combinar las rutas para obtener las rutas completas
    ruta_origen = os.getcwd()+'/output/'+original_network.name + \
        '/'+'salida_optimizacion.xlsx'
    ruta_destino = os.getcwd()+'/output/'+new_network.name + \
        '/'+'salida_optimizacion.xlsx'

    # Intentar copiar el archivo
    try:
        # shutil.copy() realiza una copia superficial del archivo
        shutil.copy(ruta_origen, ruta_destino)
        print(
            f"Archivo '{new_network.archivo}' /salida_optimizacion.xlsx copiado con éxito.")
    except FileNotFoundError:
        print(
            f"El archivo '{new_network.archivo}' /salida_optimizacion.xlsx no fue encontrado en carpeta origen.")
    except PermissionError:
        print(
            f"No tienes permisos para copiar el archivo {new_network.name} /salida_optimizacion.xlsx .")
    except Exception as e:
        print(f"Error inesperado: {e}")

    read_data.read_parameters(new_network)
    read_data.read_file_excel(new_network, path)
    read_data.delete_surplus_data(new_network)
    read_data.merge_niveles_capac(new_network, post_optima=True)
    read_data.create_df_asignacion(new_network, post_optima=True)
    read_data.create_df_probs_kk(new_network)
    read_data.create_df_arcos(new_network, post_optima=True)


# %% <codecell> main
if __name__ == "__main__":

        # Borro carpeta con resultados previos
        
        data_functions.borrar_contenido_carpeta(os.getcwd()+'/output/')
        #print("\nContenidos borrados. \nContinuando...")
        
        networks_dict={} #Diccionario con las redes utilizadas en el programa
        problems_dict={} #Diccionario con los problemas y las soluciones a la red del programa

        # Indicamos origen de datos y definimos valores I,J,K
        I,J,K= [6,6,6]
        archivo = r"C:\Users\edgar\OneDrive - Universidad Libre\Doctorado\Códigos Python\HcNDP\Health-Care-Network-Design-Problem\hcndp/data/red_original/datos_i16_j10_k10_base.xlsx"
        # Objeto network

        # Creamos un objeto network
        _name="red_original"

        networks_dict[_name] = network.Network(I,J,K,archivo,_name)
        networks_dict[_name].create_folders()
        #print (f"\nSe ha creado exitosamente el objeto {_name}.")


        # Llenar objeto con datos de Excel

        networks_dict[_name].read_file_excel('C:/Users/edgar/OneDrive - Universidad Libre/Doctorado/Códigos Python/HcNDP/Health-Care-Network-Design-Problem/data/red_original/datos_i16_j10_k10_base.xlsx')
        networks_dict[_name].delete_surplus_data()

        #print ("#" * 60)
        #print (f"\nSe han cargado exitosamente los datos en el objeto {_name}.")
    

        # Creo el objeto solucion

        solutions.create_problem_object(networks_dict['red_original'], problems_dict, name_problem="temporal")
        current_solution = problems_dict["temporal"]

        # Defino objetivo y método
        current_solution.optimizar=True
        current_solution.tecnica="Local_Search"
        _objective_and_description =['1', 'Minimizar congestión máxima (rho)']
        current_solution.objective = _objective_and_description[0]
        current_solution.description_objective = _objective_and_description[1]
        current_solution.name_problem = _objective_and_description[1]+" "+current_solution.tecnica
        
        # Actualizo nombre de la solución en solution_dict (Ya no es "temporal")
        _clave_temporal = 'temporal'
        _solucion_temporal = problems_dict[_clave_temporal]
        
        problems_dict[_solucion_temporal.name_problem] = problems_dict.pop(_clave_temporal)
        problems_dict[_solucion_temporal.name_problem].network_copy.name_problem = \
            problems_dict[_solucion_temporal.name_problem].name_problem
        print (f"Se ha actualizado el objeto {problems_dict[_solucion_temporal.name_problem].name_problem}")
        
        # Ejecuto initial solution
        current_solution= current_solution.initial_solution(networks_dict['red_original'])
        local_search.calcular_kpi_local_search(current_solution)
        
        current_solution.fix_initial_solution()
                    
    