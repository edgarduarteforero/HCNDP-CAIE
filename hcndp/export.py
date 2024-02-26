# -*- coding: utf-8 -*-
"""
Created on Tue Dec 19 18:29:12 2023

@author: edgar
"""


           
def export_data(network):
    import pandas as pd
    import os  

    output_file=os.getcwd()+'/output/'+network.name_problem+'/salida_medicion.xlsx'
        
    with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
        # Iterar a través del diccionario y escribir cada DataFrame en una hoja
        for sheet_name, df in network.file.items():
            if isinstance(df, pd.DataFrame):
                df.to_excel(writer, sheet_name=sheet_name, index=True)

def create_index_sheet(network):
    # TODO Personalizar el texto que explica cada hoja. 
    # Específicamente la variable sheet_ref
    
    #Creo una hoja con índices en el archivo de excel
    import openpyxl
    import os
    from openpyxl import Workbook
    #from openpyxl.styles import Hyperlink
    
    from openpyxl.styles import NamedStyle
    from openpyxl.styles import NamedStyle, Font, colors

    hyperlink_style = NamedStyle(name='Hyperlink', font=Font(color=colors.BLUE, underline='single'))


    
    # Nombre del archivo Excel
    excel_file = os.getcwd()+'/output/'+network.name_problem+'/salida_medicion.xlsx'
    
    # Cargar el archivo existente
    workbook = openpyxl.load_workbook(excel_file)
    
    # Crear una nueva hoja para el índice
    index_sheet = workbook.create_sheet('Índice', 0)
    
    # Encabezado del índice
    index_sheet['A1'] = 'Hoja'
    index_sheet['B1'] = 'Referencia'
    index_sheet['C1'] = 'Enlace'
    
    # Obtener una lista de nombres de hojas en el archivo
    sheet_names = workbook.sheetnames
    
    # Llenar el índice con nombres de hojas, referencias y enlaces
    for row, sheet_name in enumerate(sheet_names, start=2):
        sheet_ref = f"'{sheet_name}'!A1"  # Referencia de celda en la hoja
        sheet_link = f'#{sheet_name}!A1'  # Enlace a la hoja
        index_sheet.cell(row=row, column=1, value=sheet_name)
        index_sheet.cell(row=row, column=2, value=sheet_ref)
        #index_sheet.cell(row=row, column=3, value=sheet_link)
        index_sheet.cell(row=row, column=3, value=sheet_link).style = hyperlink_style
        index_sheet.cell(row=row, column=3).hyperlink = sheet_link
        
    # Guardar el archivo actualizado
    workbook.save(excel_file)



